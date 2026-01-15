from __future__ import annotations

"""Excel logger for the LaTeX SolarGeoRisk EPEC runs.

Design goals
- Robust: never crashes on missing keys in `hist` rows.
- Transparent: 1 row = (iteration, player) snapshot.
- Useful: logs strategic vars, primal flows, lambda, and convergence metrics if provided.

Typical `hist` row (recommended)
{
  'iter': int,
  'region': str,
  'accepted': bool,
  'status': str,
  'term': str,
  'msg': str,
  'ulp_obj': float,
  'lambda': float,
  # optional iteration metrics (repeat on each row or only on summary rows)
  'abs_max_change': float,
  'rel_inf': float,

  # strategic vars (either as dicts for all regions or scalars for just the player)
  'q_man': {r: val}, 'd_mod': {r: val}, 'sigma': {r: val}, 'beta': {r: val},

  # primal vars
  'x_man': {r: val}, 'x_dem': {r: val},
  'x_mod': {(e,r): val},
}

If your `hist` stores only the player's values (scalars), the writer will still log
those into the '*_player' columns.
"""

from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as e:  # pragma: no cover
    raise ModuleNotFoundError(
        "openpyxl is required to write Excel results. Install with: pip install openpyxl"
    ) from e


def _safe_value(x: Any) -> Any:
    """Convert Pyomo/numpy scalars to plain python values when possible."""
    try:
        if hasattr(x, "value"):
            x = x.value
        if hasattr(x, "item") and callable(x.item):
            x = x.item()
        if isinstance(x, (int, float)):
            return float(x)
        return x
    except Exception:
        return x


def _auto_fit_columns(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
        values = ["" if v is None else str(v) for v in col]
        width = min(max((len(v) for v in values), default=0) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)


def _flatten_region(prefix: str, regions: Iterable[str], d: Optional[Mapping[str, Any]]) -> Dict[str, Any]:
    d = d or {}
    return {f"{prefix}_{r}": _safe_value(d.get(r)) for r in regions}


def _flatten_arcs(prefix: str, arcs: Iterable[Tuple[str, str]], d: Optional[Mapping[Tuple[str, str], Any]]) -> Dict[str, Any]:
    d = d or {}
    out: Dict[str, Any] = {}
    for (e, r) in arcs:
        out[f"{prefix}_{e}_{r}"] = _safe_value(d.get((e, r)))
    return out


def _theta_dict(theta: Any) -> Dict[str, Any]:
    """Best-effort extraction for final theta sheet."""
    if theta is None:
        return {}
    if is_dataclass(theta):
        return asdict(theta)
    out: Dict[str, Any] = {}
    for k in ("q_man", "d_mod", "sigma", "beta"):
        if hasattr(theta, k):
            out[k] = getattr(theta, k)
    return out


def save_run_results_excel(
    *,
    project_root: Path,
    sets: Any,
    params: Any,
    theta_star: Any,
    hist: List[Dict[str, Any]],
    run_cfg: Optional[Dict[str, Any]] = None,
    solver_opts: Optional[Dict[str, Any]] = None,
    filename_prefix: str = "run_latex",
    results_subdir: str = "results",
) -> Path:
    """Write an Excel file with run history.

    - Output path: <project_root>/<results_subdir>/<prefix>_<timestamp>.xlsx
    - Sheet 'history_wide': one row per hist entry (iter, region)
    - Sheet 'theta_final': final strategic variables
    - Sheet 'run_config': run_cfg + solver options + selected params summary
    """

    project_root = Path(project_root)
    results_dir = project_root / results_subdir
    results_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = results_dir / f"{filename_prefix}_{ts}.xlsx"

    R = list(getattr(sets, "R"))
    # Use A if present; else fall back to RR; else empty
    A = list(getattr(sets, "A", getattr(sets, "RR", [])))

    wb = Workbook()

    # ----------------
    # run_config sheet
    # ----------------
    ws_cfg = wb.active
    ws_cfg.title = "run_config"
    ws_cfg.append(["key", "value"])
    ws_cfg["A1"].font = Font(bold=True)
    ws_cfg["B1"].font = Font(bold=True)

    def _append_kv(prefix: str, d: Optional[Dict[str, Any]]):
        if not d:
            return
        for k in sorted(d.keys()):
            ws_cfg.append([f"{prefix}{k}", _safe_value(d[k])])

    _append_kv("run.", run_cfg or {})
    _append_kv("solver.", solver_opts or {})

    # minimal param summary (avoid huge matrices)
    # log caps + core costs if available
    for name in ("Qcap", "Dcap", "c_man", "beta_bar"):
        if hasattr(params, name):
            d = getattr(params, name)
            if isinstance(d, dict):
                for r in R:
                    ws_cfg.append([f"param.{name}.{r}", _safe_value(d.get(r))])

    ws_cfg.freeze_panes = "A2"
    _auto_fit_columns(ws_cfg)

    # ----------------
    # theta_final sheet
    # ----------------
    ws_t = wb.create_sheet("theta_final")
    ws_t.append(["var", "key", "value"])
    for c in ("A1", "B1", "C1"):
        ws_t[c].font = Font(bold=True)

    th = _theta_dict(theta_star)
    for var in ("q_man", "d_mod", "sigma", "beta"):
        d = th.get(var, {}) if isinstance(th, dict) else {}
        if isinstance(d, dict):
            for r in R:
                ws_t.append([var, r, _safe_value(d.get(r))])

    ws_t.freeze_panes = "A2"
    _auto_fit_columns(ws_t)

    # ----------------
    # history_wide sheet
    # ----------------
    ws = wb.create_sheet("history_wide")

    # Build columns
    base_cols = [
        "iter",
        "region",
        "accepted",
        "status",
        "term",
        "msg",
        "ulp_obj",
        "lambda",
        "abs_max_change",
        "rel_inf",
    ]

    # Wide: strategic variables (all regions)
    strat_cols: List[str] = []
    for pfx in ("q_man", "d_mod", "sigma", "beta"):
        strat_cols.extend([f"{pfx}_{r}" for r in R])
        strat_cols.append(f"{pfx}_player")

    # Wide: primal vars
    primal_cols: List[str] = []
    for pfx in ("x_man", "x_dem"):
        primal_cols.extend([f"{pfx}_{r}" for r in R])

    arc_cols: List[str] = [f"x_mod_{e}_{r}" for (e, r) in A]

    cols = base_cols + strat_cols + primal_cols + arc_cols

    ws.append(cols)
    for j in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Fill rows
    for row in hist:
        flat: Dict[str, Any] = {}

        # base
        for k in base_cols:
            flat[k] = _safe_value(row.get(k))

        # allow alternative names
        if flat.get("lambda") is None:
            flat["lambda"] = _safe_value(row.get("lam"))

        # strategic dicts (all regions)
        for pfx in ("q_man", "d_mod", "sigma", "beta"):
            d = row.get(pfx)
            if isinstance(d, dict):
                flat.update(_flatten_region(pfx, R, d))
            else:
                # no dict: leave region columns blank
                flat.update({f"{pfx}_{r}": None for r in R})

            # player scalar (preferred key br_*, fallback to var in model)
            player_key = f"br_{pfx}" if f"br_{pfx}" in row else None
            if player_key:
                flat[f"{pfx}_player"] = _safe_value(row.get(player_key))
            else:
                # if it is a scalar already
                flat[f"{pfx}_player"] = _safe_value(d) if not isinstance(d, dict) else None

        # primal dicts
        for pfx in ("x_man", "x_dem"):
            d = row.get(pfx)
            if isinstance(d, dict):
                flat.update(_flatten_region(pfx, R, d))
            else:
                flat.update({f"{pfx}_{r}": None for r in R})

        # arc flows
        xmod = row.get("x_mod")
        if isinstance(xmod, dict):
            flat.update(_flatten_arcs("x_mod", A, xmod))
        else:
            # empty arc flow columns
            for (e, r) in A:
                flat[f"x_mod_{e}_{r}"] = None

        ws.append([flat.get(c) for c in cols])

    _auto_fit_columns(ws)

    wb.save(out_path)
    return out_path
