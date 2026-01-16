from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .results import CaseData, Theta, LLPResult, GSHistory


SHEET_RUNCFG = "run_config"
SHEET_FINAL = "final_theta"
SHEET_HIST = "history"


def _utc_stamp() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")


def _ensure_wb(path: Path) -> Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    # remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    return wb


def _get_or_create_sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _is_sheet_empty(ws: Worksheet) -> bool:
    return ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None


def _write_header_if_needed(ws: Worksheet, header: Sequence[str]) -> None:
    if _is_sheet_empty(ws):
        ws.append(list(header))
        ws.freeze_panes = "A2"
        return

    # If header differs, create a new sheet variant
    existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing = [x for x in existing if x is not None]
    if existing and list(existing) != list(header):
        # create new sheet with suffix
        base = ws.title
        wb = ws.parent
        k = 2
        while f"{base}_{k}" in wb.sheetnames:
            k += 1
        ws2 = wb.create_sheet(f"{base}_{k}")
        ws2.append(list(header))
        ws2.freeze_panes = "A2"


def _append_kv_block(ws: Worksheet, title: str, kv: Dict[str, Any]) -> None:
    # Append a run block at the end
    if _is_sheet_empty(ws):
        ws.append(["key", "value"])

    ws.append([None, None])
    ws.append([f"--- {title} ---", _utc_stamp()])
    for k, v in kv.items():
        ws.append([str(k), v])


def _theta_to_rows(theta: Theta, regions: List[str]) -> List[Tuple[str, str, float]]:
    rows: List[Tuple[str, str, float]] = []
    for cat, d in [
        ("q_man", theta.q_man),
        ("d_mod", theta.d_mod),
        ("sigma", theta.sigma),
        ("beta", theta.beta),
    ]:
        for r in regions:
            rows.append((cat, r, float(d.get(r, 0.0))))
    return rows


def _compute_llp_obj(data: CaseData, theta: Theta, llp: LLPResult) -> float:
    # LLP objective:
    # min Σ_r sigma[r]*x_man[r] + Σ_{e,r} c_ship[e,r]*x_mod[e,r] - Σ_r beta[r]*x_dem[r]
    val = 0.0
    for r in data.regions:
        val += float(theta.sigma[r]) * float(llp.x_man.get(r, 0.0))
        val -= float(theta.beta[r]) * float(llp.x_dem.get(r, 0.0))
    for (e, r), x in llp.x_mod.items():
        val += float(data.c_ship[(e, r)]) * float(x)
    return float(val)


def _pack_step_row(
    data: CaseData,
    row: Dict[str, Any],
    header: Sequence[str],
) -> List[Any]:
    # Missing keys -> None
    return [row.get(k, None) for k in header]


def default_history_header(data: CaseData) -> List[str]:
    R = data.regions
    header = [
        "run_stamp",
        "iter",
        "region",
        "accepted",
        "method",
        "model_status",
        "solve_status",
        "solver_objective",
        "ulp_obj",
        "llp_obj",
        "lambda",
        "abs_max_change",
        "rel_max_change",
        "elapsed_s",
    ]

    # full theta snapshot after update
    for r in R:
        header.append(f"q_man_{r}")
    for r in R:
        header.append(f"d_mod_{r}")
    for r in R:
        header.append(f"sigma_{r}")
    for r in R:
        header.append(f"beta_{r}")

    # player BR (pre-damping) for this step
    header += ["br_q_man", "br_d_mod", "br_sigma", "br_beta"]

    # LLP primals
    for r in R:
        header.append(f"x_man_{r}")
    for r in R:
        header.append(f"x_dem_{r}")
    for e in R:
        for r in R:
            header.append(f"x_mod_{e}_{r}")

    # Optional: selected duals (useful debugging)
    for r in R:
        header.append(f"pi_{r}")
    for r in R:
        header.append(f"mu_{r}")
    for r in R:
        header.append(f"alpha_{r}")
    for r in R:
        header.append(f"phi_{r}")

    return header


def write_gs_results_excel(
    path: str | Path,
    data: CaseData,
    run_config: Dict[str, Any],
    theta_final: Theta,
    history_rows: List[Dict[str, Any]],
    append: bool = True,
) -> Path:
    """
    Writes/updates an Excel workbook with:
      - run_config (key/value blocks appended)
      - final_theta (category/key/value blocks appended)
      - history (tabular, appended rows)

    If workbook exists and `append=True`, rows are appended.
    """
    path = Path(path)
    wb = _ensure_wb(path)

    ws_cfg = _get_or_create_sheet(wb, SHEET_RUNCFG)
    ws_final = _get_or_create_sheet(wb, SHEET_FINAL)
    ws_hist = _get_or_create_sheet(wb, SHEET_HIST)

    # run_config block
    _append_kv_block(ws_cfg, "run_config", run_config)

    # final_theta block
    if _is_sheet_empty(ws_final):
        ws_final.append(["category", "key", "value"])
        ws_final.freeze_panes = "A2"
    ws_final.append([None, None, None])
    ws_final.append([f"--- final_theta ---", _utc_stamp(), None])
    for cat, k, v in _theta_to_rows(theta_final, data.regions):
        ws_final.append([cat, k, v])

    # history table
    header = default_history_header(data)
    _write_header_if_needed(ws_hist, header)

    # If header mismatch caused a new sheet, use last one
    ws_hist = wb[wb.sheetnames[-1]] if wb.sheetnames[-1].startswith(SHEET_HIST) else ws_hist

    for row in history_rows:
        ws_hist.append(_pack_step_row(data, row, header))

    wb.save(path)
    return path
